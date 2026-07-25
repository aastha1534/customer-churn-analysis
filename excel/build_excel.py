"""
build_excel.py
---------------
Builds Customer_Churn_Analysis.xlsx with:
  - RawData sheet: full dataset
  - Summary sheet: KPI cards + churn-by-segment tables, all driven by
    live SUMIFS/COUNTIFS/AVERAGEIFS formulas (recalculates if RawData changes)

Run:
    python build_excel.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
KPI_LABEL_FONT = Font(name=FONT_NAME, size=10, color="595959")
KPI_VALUE_FONT = Font(name=FONT_NAME, bold=True, size=20, color="1F4E78")
BODY_FONT = Font(name=FONT_NAME, size=10)
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

df = pd.read_csv("../data/telco_customer_churn.csv")

wb = Workbook()

# ---------------------------------------------------------------- RAW DATA
ws_raw = wb.active
ws_raw.title = "RawData"
for r in dataframe_to_rows(df, index=False, header=True):
    ws_raw.append(r)
for c in range(1, df.shape[1] + 1):
    cell = ws_raw.cell(row=1, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ws_raw.column_dimensions[get_column_letter(c)].width = 16
ws_raw.freeze_panes = "A2"
n_rows = df.shape[0] + 1  # + header

# Column letters for reference
col_letter = {name: get_column_letter(i + 1) for i, name in enumerate(df.columns)}
churn_rng = f"RawData!${col_letter['Churn']}$2:${col_letter['Churn']}$_N_".replace("_N_", str(n_rows))
contract_rng = f"RawData!${col_letter['Contract']}$2:${col_letter['Contract']}$_N_".replace("_N_", str(n_rows))
internet_rng = f"RawData!${col_letter['InternetService']}$2:${col_letter['InternetService']}$_N_".replace("_N_", str(n_rows))
payment_rng = f"RawData!${col_letter['PaymentMethod']}$2:${col_letter['PaymentMethod']}$_N_".replace("_N_", str(n_rows))
monthly_rng = f"RawData!${col_letter['MonthlyCharges']}$2:${col_letter['MonthlyCharges']}$_N_".replace("_N_", str(n_rows))
tenure_rng = f"RawData!${col_letter['tenure']}$2:${col_letter['tenure']}$_N_".replace("_N_", str(n_rows))

# ---------------------------------------------------------------- SUMMARY
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c in "BCDEFGH":
    ws.column_dimensions[c].width = 18

ws["B2"] = "Telco Customer Churn — KPI Dashboard"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Source: RawData sheet (7,043 customers) — all figures are live formulas"
ws["B3"].font = Font(name=FONT_NAME, italic=True, size=9, color="808080")

# ---- KPI cards row (B5:H9) ----
kpis = [
    ("Total Customers", f"=COUNTA({churn_rng})", "#,##0"),
    ("Churned Customers", f"=COUNTIF({churn_rng},\"Yes\")", "#,##0"),
    ("Churn Rate", f"=COUNTIF({churn_rng},\"Yes\")/COUNTA({churn_rng})", "0.0%"),
    ("Avg Monthly Charges", f"=AVERAGE({monthly_rng})", "$#,##0.00"),
    ("Monthly Revenue at Risk", f"=SUMIF({churn_rng},\"Yes\",{monthly_rng})", "$#,##0"),
    ("Avg Tenure (Churned)", f"=AVERAGEIF({churn_rng},\"Yes\",{tenure_rng})", "0.0"),
]
col_start = 2  # column B
for i, (label, formula, fmt) in enumerate(kpis):
    col = col_start + i
    letter = get_column_letter(col)
    ws.cell(row=5, column=col, value=label).font = KPI_LABEL_FONT
    ws.cell(row=5, column=col).alignment = Alignment(wrap_text=True)
    vcell = ws.cell(row=6, column=col, value=formula)
    vcell.font = KPI_VALUE_FONT
    vcell.number_format = fmt
    for row in (5, 6):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="EAF1F8")
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[6].height = 30

# ---- Churn by Contract table ----
start_row = 9
ws.cell(row=start_row, column=2, value="Churn Rate by Contract Type").font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
headers = ["Contract", "Customers", "Churned", "Churn Rate"]
for i, h in enumerate(headers):
    cell = ws.cell(row=start_row + 1, column=2 + i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

contracts = ["Month-to-month", "One year", "Two year"]
for i, contract in enumerate(contracts):
    r = start_row + 2 + i
    ws.cell(row=r, column=2, value=contract).font = BODY_FONT
    ws.cell(row=r, column=3, value=f'=COUNTIF({contract_rng},"{contract}")').font = BODY_FONT
    ws.cell(row=r, column=4, value=f'=COUNTIFS({contract_rng},"{contract}",{churn_rng},"Yes")').font = BODY_FONT
    cell = ws.cell(row=r, column=5, value=f"=D{r}/C{r}")
    cell.number_format = "0.0%"
    cell.font = BODY_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c).border = BORDER

# ---- Churn by Internet Service table ----
start_row2 = start_row + 2 + len(contracts) + 2
ws.cell(row=start_row2, column=2, value="Churn Rate by Internet Service").font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
for i, h in enumerate(headers):
    cell = ws.cell(row=start_row2 + 1, column=2 + i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

services = ["DSL", "Fiber optic", "No"]
for i, svc in enumerate(services):
    r = start_row2 + 2 + i
    ws.cell(row=r, column=2, value=svc).font = BODY_FONT
    ws.cell(row=r, column=3, value=f'=COUNTIF({internet_rng},"{svc}")').font = BODY_FONT
    ws.cell(row=r, column=4, value=f'=COUNTIFS({internet_rng},"{svc}",{churn_rng},"Yes")').font = BODY_FONT
    cell = ws.cell(row=r, column=5, value=f"=D{r}/C{r}")
    cell.number_format = "0.0%"
    cell.font = BODY_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c).border = BORDER

# ---- Churn by Payment Method table ----
start_row3 = start_row2 + 2 + len(services) + 2
ws.cell(row=start_row3, column=2, value="Churn Rate by Payment Method").font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
for i, h in enumerate(headers):
    cell = ws.cell(row=start_row3 + 1, column=2 + i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

payments = ["Electronic check", "Mailed check", "Bank transfer (automatic)", "Credit card (automatic)"]
for i, pm in enumerate(payments):
    r = start_row3 + 2 + i
    ws.cell(row=r, column=2, value=pm).font = BODY_FONT
    ws.cell(row=r, column=3, value=f'=COUNTIF({payment_rng},"{pm}")').font = BODY_FONT
    ws.cell(row=r, column=4, value=f'=COUNTIFS({payment_rng},"{pm}",{churn_rng},"Yes")').font = BODY_FONT
    cell = ws.cell(row=r, column=5, value=f"=D{r}/C{r}")
    cell.number_format = "0.0%"
    cell.font = BODY_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c).border = BORDER

# Note documenting data source
note_row = start_row3 + 2 + len(payments) + 2
ws.cell(row=note_row, column=2,
        value="Note: All figures calculated live via COUNTIF/COUNTIFS/AVERAGEIF/SUMIF formulas "
              "referencing the RawData sheet. Source: synthetic dataset generated to match the "
              "public IBM Telco Customer Churn schema (see /python/generate_data.py).")
ws.cell(row=note_row, column=2).font = Font(name=FONT_NAME, italic=True, size=8, color="808080")

wb.save("Customer_Churn_Analysis.xlsx")
print("Saved Customer_Churn_Analysis.xlsx")
